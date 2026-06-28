"""Parse uploaded files of any format into flat text rows for indexing.

Supported formats: CSV, TSV, JSON, XML, Excel (.xlsx/.xls), TXT.
"""

import csv
import io
import json
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


def parse_csv(content: bytes, encoding: str = "utf-8") -> list[str]:
    text = content.decode(encoding)
    reader = csv.DictReader(io.StringIO(text))
    rows: list[str] = []
    for record in reader:
        row_text = " | ".join(f"{k}: {v}" for k, v in record.items() if v)
        if row_text:
            rows.append(row_text)
    return rows


def parse_tsv(content: bytes, encoding: str = "utf-8") -> list[str]:
    text = content.decode(encoding)
    reader = csv.DictReader(io.StringIO(text), delimiter="\t")
    rows: list[str] = []
    for record in reader:
        row_text = " | ".join(f"{k}: {v}" for k, v in record.items() if v)
        if row_text:
            rows.append(row_text)
    return rows


def parse_json(content: bytes, encoding: str = "utf-8") -> list[str]:
    text = content.decode(encoding)
    data = json.loads(text)

    if isinstance(data, list):
        items = data
    elif isinstance(data, dict):
        items = [data]
    else:
        raise ValueError(f"Unsupported JSON root type: {type(data).__name__}")

    rows: list[str] = []
    for item in items:
        if isinstance(item, dict):
            row_text = " | ".join(f"{k}: {v}" for k, v in item.items() if v is not None)
        else:
            row_text = str(item)
        if row_text:
            rows.append(row_text)
    return rows


def parse_xml(content: bytes) -> list[str]:
    root = ET.fromstring(content)
    rows: list[str] = []
    for element in root:
        parts: list[str] = []
        if element.text and element.text.strip():
            parts.append(f"{element.tag}: {element.text.strip()}")
        for child in element:
            text = child.text.strip() if child.text else ""
            if text:
                parts.append(f"{child.tag}: {text}")
            for attr_key, attr_val in child.attrib.items():
                parts.append(f"{attr_key}: {attr_val}")
        for attr_key, attr_val in element.attrib.items():
            parts.append(f"{attr_key}: {attr_val}")
        if parts:
            rows.append(" | ".join(parts))
    return rows


def parse_excel(content: bytes) -> list[str]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    rows: list[str] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_rows = list(ws.iter_rows(values_only=True))
        if not sheet_rows:
            continue
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(sheet_rows[0])]
        for data_row in sheet_rows[1:]:
            parts = []
            for header, val in zip(headers, data_row):
                if val is not None and str(val).strip():
                    parts.append(f"{header}: {val}")
            if parts:
                rows.append(" | ".join(parts))
    wb.close()
    return rows


def parse_txt(content: bytes, encoding: str = "utf-8") -> list[str]:
    text = content.decode(encoding)
    rows: list[str] = []
    for line in text.splitlines():
        stripped = line.strip()
        if stripped:
            rows.append(stripped)
    return rows


SUPPORTED_EXTENSIONS = {
    ".csv": "CSV",
    ".tsv": "TSV",
    ".json": "JSON",
    ".xml": "XML",
    ".xlsx": "Excel",
    ".xls": "Excel",
    ".txt": "Text",
}


def parse_file(filename: str, content: bytes) -> list[str]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return parse_csv(content)
    if lower.endswith(".tsv"):
        return parse_tsv(content)
    if lower.endswith(".json"):
        return parse_json(content)
    if lower.endswith(".xml"):
        return parse_xml(content)
    if lower.endswith((".xlsx", ".xls")):
        return parse_excel(content)
    if lower.endswith(".txt"):
        return parse_txt(content)
    supported = ", ".join(SUPPORTED_EXTENSIONS.keys())
    raise ValueError(f"Unsupported file type: {filename}. Supported: {supported}")
