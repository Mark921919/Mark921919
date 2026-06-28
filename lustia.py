import os
import sys
import time
import json
import csv
import io
import sqlite3
import threading
import xml.etree.ElementTree as ET
import requests
from datetime import datetime
from pathlib import Path
import configparser
import glob
from contextlib import asynccontextmanager

try:
    from tqdm import tqdm
    from colorama import init, Fore, Style
    init(autoreset=True)
except ImportError:
    print("Установи зависимости: pip install requests tqdm colorama fastapi uvicorn python-multipart openpyxl")
    sys.exit(1)

try:
    import uvicorn
    from fastapi import FastAPI, APIRouter, Form, HTTPException, Query, UploadFile
    from openpyxl import load_workbook
except ImportError:
    print("Установи зависимости: pip install fastapi uvicorn python-multipart openpyxl")
    sys.exit(1)

# ========================== КОНФИГУРАЦИЯ ==========================

SERVER_HOST = "127.0.0.1"
SERVER_PORT = 8000
BASE_URL = f"http://{SERVER_HOST}:{SERVER_PORT}/api"
API_KEY = ""

# Кеш для мгновенного поиска
cache = {}
CACHE_TTL = 300  # 5 минут

# Конфиг для загрузчика
CONFIG_FILE = Path.home() / ".dbuploader.ini"
SUPPORTED = {".csv", ".json", ".xlsx", ".xls", ".txt", ".tsv", ".xml"}
CHUNK_SIZE = 1024 * 1024  # 1 МБ
MAX_RETRIES = 3
RETRY_DELAY = 3  # сек

# ========================== ВСТРОЕННЫЙ API СЕРВЕР ==========================

_DB_DIR = Path(__file__).resolve().parent / "data"
_DB_PATH = _DB_DIR / "store.db"
_local = threading.local()


def _get_conn() -> sqlite3.Connection:
    conn = getattr(_local, "conn", None)
    if conn is None:
        _DB_DIR.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(str(_DB_PATH), check_same_thread=False)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        _local.conn = conn
    return conn


def init_db() -> None:
    conn = _get_conn()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS databases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            db_id INTEGER NOT NULL REFERENCES databases(id) ON DELETE CASCADE,
            row_data TEXT NOT NULL
        );
        CREATE VIRTUAL TABLE IF NOT EXISTS records_fts USING fts5(
            row_data, content='records', content_rowid='id'
        );
        CREATE TRIGGER IF NOT EXISTS records_ai AFTER INSERT ON records BEGIN
            INSERT INTO records_fts(rowid, row_data) VALUES (new.id, new.row_data);
        END;
        CREATE TRIGGER IF NOT EXISTS records_ad AFTER DELETE ON records BEGIN
            INSERT INTO records_fts(records_fts, rowid, row_data)
                VALUES ('delete', old.id, old.row_data);
        END;
    """)
    conn.commit()


def create_database_entry(name: str) -> int:
    conn = _get_conn()
    cur = conn.execute("INSERT INTO databases (name) VALUES (?)", (name,))
    conn.commit()
    return cur.lastrowid


def get_database_entry(name: str) -> dict | None:
    conn = _get_conn()
    row = conn.execute("SELECT * FROM databases WHERE name = ?", (name,)).fetchone()
    return dict(row) if row else None


def list_databases_db() -> list[dict]:
    conn = _get_conn()
    rows = conn.execute("SELECT * FROM databases ORDER BY id").fetchall()
    return [dict(r) for r in rows]


def delete_database_entry(db_id: int) -> None:
    conn = _get_conn()
    conn.execute("DELETE FROM records WHERE db_id = ?", (db_id,))
    conn.execute("DELETE FROM databases WHERE id = ?", (db_id,))
    conn.commit()


def insert_records_db(db_id: int, rows: list[str]) -> int:
    conn = _get_conn()
    conn.executemany(
        "INSERT INTO records (db_id, row_data) VALUES (?, ?)",
        [(db_id, r) for r in rows],
    )
    conn.commit()
    return len(rows)


def get_columns(db_id: int) -> list[str]:
    conn = _get_conn()
    row = conn.execute(
        "SELECT row_data FROM records WHERE db_id = ? LIMIT 1", (db_id,)
    ).fetchone()
    if not row:
        return []
    parts = row["row_data"].split(" | ")
    cols = []
    for p in parts:
        if ": " in p:
            cols.append(p.split(": ", 1)[0])
    return cols


def search_records_db(
    query: str, limit: int = 50, offset: int = 0, db_name: str | None = None
) -> tuple[list[dict], int]:
    conn = _get_conn()
    if db_name:
        count_sql = """
            SELECT COUNT(*) AS cnt
            FROM records_fts fts
            JOIN records r ON r.id = fts.rowid
            JOIN databases d ON d.id = r.db_id
            WHERE fts.row_data MATCH ? AND d.name = ?
        """
        total = conn.execute(count_sql, (query, db_name)).fetchone()["cnt"]
        sql = """
            SELECT r.id, d.name AS db_name, r.row_data, rank
            FROM records_fts fts
            JOIN records r ON r.id = fts.rowid
            JOIN databases d ON d.id = r.db_id
            WHERE fts.row_data MATCH ? AND d.name = ?
            ORDER BY rank LIMIT ? OFFSET ?
        """
        rows = conn.execute(sql, (query, db_name, limit, offset)).fetchall()
    else:
        count_sql = """
            SELECT COUNT(*) AS cnt
            FROM records_fts fts
            JOIN records r ON r.id = fts.rowid
            WHERE fts.row_data MATCH ?
        """
        total = conn.execute(count_sql, (query,)).fetchone()["cnt"]
        sql = """
            SELECT r.id, d.name AS db_name, r.row_data, rank
            FROM records_fts fts
            JOIN records r ON r.id = fts.rowid
            JOIN databases d ON d.id = r.db_id
            WHERE fts.row_data MATCH ?
            ORDER BY rank LIMIT ? OFFSET ?
        """
        rows = conn.execute(sql, (query, limit, offset)).fetchall()

    results = []
    for r in rows:
        item: dict = {"database": r["db_name"]}
        for part in r["row_data"].split(" | "):
            if ": " in part:
                k, v = part.split(": ", 1)
                item[k] = v
            else:
                item["data"] = part
        results.append(item)
    return results, total


def count_records_db(db_id: int) -> int:
    conn = _get_conn()
    row = conn.execute(
        "SELECT COUNT(*) AS cnt FROM records WHERE db_id = ?", (db_id,)
    ).fetchone()
    return row["cnt"]


# ========================== ПАРСЕР ФАЙЛОВ ==========================

_ENCODINGS = ["utf-8", "utf-8-sig", "cp1251", "latin-1"]


def _decode(content: bytes) -> str:
    for enc in _ENCODINGS:
        try:
            return content.decode(enc)
        except (UnicodeDecodeError, UnicodeError):
            continue
    return content.decode("latin-1")


def parse_csv(content: bytes) -> list[str]:
    text = _decode(content)
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for record in reader:
        row_text = " | ".join(f"{k}: {v}" for k, v in record.items() if v)
        if row_text:
            rows.append(row_text)
    return rows


def parse_tsv(content: bytes) -> list[str]:
    text = _decode(content)
    reader = csv.DictReader(io.StringIO(text), delimiter="\t")
    rows = []
    for record in reader:
        row_text = " | ".join(f"{k}: {v}" for k, v in record.items() if v)
        if row_text:
            rows.append(row_text)
    return rows


def parse_json_file(content: bytes) -> list[str]:
    text = _decode(content)
    data = json.loads(text)
    if isinstance(data, list):
        items = data
    elif isinstance(data, dict):
        items = [data]
    else:
        raise ValueError(f"Unsupported JSON root type: {type(data).__name__}")
    rows = []
    for item in items:
        if isinstance(item, dict):
            row_text = " | ".join(
                f"{k}: {v}" for k, v in item.items() if v is not None
            )
        else:
            row_text = str(item)
        if row_text:
            rows.append(row_text)
    return rows


def parse_xml(content: bytes) -> list[str]:
    root = ET.fromstring(content)
    rows = []
    for element in root:
        parts = []
        if element.text and element.text.strip():
            parts.append(f"{element.tag}: {element.text.strip()}")
        for child in element:
            text = child.text.strip() if child.text else ""
            if text:
                parts.append(f"{child.tag}: {text}")
            for attr_key, attr_val in child.attrib.items():
                parts.append(f"{attr_key}: {attr_val}")
        for attr_key, attr_val in element.attrib.items():
            parts.append(f"{attr_key}: {attr_val}")
        if parts:
            rows.append(" | ".join(parts))
    return rows


def parse_excel(content: bytes) -> list[str]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_rows = list(ws.iter_rows(values_only=True))
        if not sheet_rows:
            continue
        headers = [
            str(h) if h is not None else f"col_{i}"
            for i, h in enumerate(sheet_rows[0])
        ]
        for data_row in sheet_rows[1:]:
            parts = []
            for header, val in zip(headers, data_row):
                if val is not None and str(val).strip():
                    parts.append(f"{header}: {val}")
            if parts:
                rows.append(" | ".join(parts))
    wb.close()
    return rows


def parse_txt(content: bytes) -> list[str]:
    text = _decode(content)
    rows = []
    for line in text.splitlines():
        stripped = line.strip()
        if stripped:
            rows.append(stripped)
    return rows


SUPPORTED_EXTENSIONS = {
    ".csv": "CSV", ".tsv": "TSV", ".json": "JSON",
    ".xml": "XML", ".xlsx": "Excel", ".xls": "Excel", ".txt": "Text",
}


def parse_file(filename: str, content: bytes) -> list[str]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return parse_csv(content)
    if lower.endswith(".tsv"):
        return parse_tsv(content)
    if lower.endswith(".json"):
        return parse_json_file(content)
    if lower.endswith(".xml"):
        return parse_xml(content)
    if lower.endswith((".xlsx", ".xls")):
        return parse_excel(content)
    if lower.endswith(".txt"):
        return parse_txt(content)
    supported = ", ".join(SUPPORTED_EXTENSIONS.keys())
    raise ValueError(f"Unsupported file type: {filename}. Supported: {supported}")


# ========================== FASTAPI МАРШРУТЫ ==========================

router = APIRouter()


@router.post("/upload")
async def api_upload(file: UploadFile, name: str = Form(...)):
    start = time.perf_counter()
    if not file.filename:
        raise HTTPException(status_code=400, detail="File has no name")
    db_name = name.strip()[:100]
    if not db_name:
        raise HTTPException(status_code=400, detail="Empty database name")
    existing = get_database_entry(db_name)
    if existing:
        raise HTTPException(status_code=409, detail=f"Database '{db_name}' already exists")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")
    try:
        rows = parse_file(file.filename, content)
    except (ValueError, UnicodeDecodeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if not rows:
        raise HTTPException(status_code=400, detail="File contains no records")
    db_id = create_database_entry(db_name)
    inserted = insert_records_db(db_id, rows)
    elapsed = time.perf_counter() - start
    columns = get_columns(db_id)
    return {
        "ok": True,
        "database": {
            "name": db_name,
            "rowCount": inserted,
            "columns": columns,
            "createdAt": get_database_entry(db_name)["created_at"],
        },
        "elapsed_seconds": round(elapsed, 4),
    }


@router.get("/databases")
async def api_list_databases():
    dbs = list_databases_db()
    result = []
    for db in dbs:
        cnt = count_records_db(db["id"])
        columns = get_columns(db["id"])
        result.append({
            "name": db["name"],
            "rowCount": cnt,
            "columns": columns,
            "createdAt": db["created_at"],
        })
    return {"databases": result}


@router.get("/databases/{db_name}")
async def api_get_database(db_name: str):
    entry = get_database_entry(db_name)
    if not entry:
        raise HTTPException(status_code=404, detail=f"Database '{db_name}' not found")
    cnt = count_records_db(entry["id"])
    columns = get_columns(entry["id"])
    return {
        "name": entry["name"],
        "rowCount": cnt,
        "columns": columns,
        "createdAt": entry["created_at"],
    }


@router.delete("/databases/{db_name}")
async def api_delete_database(db_name: str):
    entry = get_database_entry(db_name)
    if not entry:
        raise HTTPException(status_code=404, detail=f"Database '{db_name}' not found")
    delete_database_entry(entry["id"])
    return {"ok": True, "deleted": db_name}


@router.get("/formats")
async def api_formats():
    return {"formats": SUPPORTED_EXTENSIONS}


@router.get("/search")
async def api_search(
    q: str = Query(..., min_length=1),
    db: str | None = Query(None),
    limit: int = Query(100, ge=1, le=10000),
    offset: int = Query(0, ge=0),
):
    start = time.perf_counter()
    if db:
        entry = get_database_entry(db)
        if not entry:
            raise HTTPException(status_code=404, detail=f"Database '{db}' not found")
    try:
        results, total_count = search_records_db(q, limit=limit, offset=offset, db_name=db)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Search error: {exc}")
    elapsed_ms = round((time.perf_counter() - start) * 1000, 2)
    return {
        "ok": True,
        "results": results,
        "count": total_count,
        "totalCount": total_count,
        "tookMs": elapsed_ms,
    }


@asynccontextmanager
async def lifespan(a: FastAPI):
    init_db()
    yield


api_app = FastAPI(
    title="Lustia API Search",
    description="Lustia API — загрузка и поиск по базам данных любого формата",
    version="1.0.0",
    lifespan=lifespan,
)
api_app.include_router(router, prefix="/api")


@api_app.get("/health")
async def health():
    return {"status": "ok"}


def _start_server():
    """Запуск API сервера в фоновом потоке."""
    import logging
    logging.getLogger("uvicorn.error").setLevel(logging.WARNING)
    logging.getLogger("uvicorn.access").setLevel(logging.WARNING)
    config = uvicorn.Config(
        api_app,
        host=SERVER_HOST,
        port=SERVER_PORT,
        log_level="warning",
    )
    server = uvicorn.Server(config)
    server.run()


def ensure_server():
    """Запускает сервер в фоне, если он ещё не запущен."""
    try:
        r = requests.get(f"http://{SERVER_HOST}:{SERVER_PORT}/health", timeout=2)
        if r.status_code == 200:
            return True
    except Exception:
        pass

    t = threading.Thread(target=_start_server, daemon=True)
    t.start()

    for _ in range(30):
        time.sleep(0.3)
        try:
            r = requests.get(f"http://{SERVER_HOST}:{SERVER_PORT}/health", timeout=2)
            if r.status_code == 200:
                return True
        except Exception:
            pass
    return False


# ========================== ЦВЕТА ==========================

class Colors:
    RED = '\033[91m'
    DARK_RED = '\033[31m'
    BLOOD_RED = '\033[38;5;88m'
    WHITE = '\033[97m'
    GRAY = '\033[90m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    CYAN = '\033[96m'
    MAGENTA = '\033[95m'
    BOLD = '\033[1m'
    END = '\033[0m'

def c(text, color=Colors.WHITE):
    return f"{color}{text}{Colors.END}"

def clear():
    os.system('cls' if os.name == 'nt' else 'clear')

# Цветные сообщения для загрузчика
def ok(msg): print(Fore.GREEN + "✅ " + msg)
def err(msg): print(Fore.RED + "❌ " + msg)
def info(msg): print(Fore.CYAN + "ℹ " + msg)
def warn(msg): print(Fore.YELLOW + "⚠ " + msg)

def hr():
    print(Fore.RED + Style.DIM + "─" * 70)

def fmt_size(n: int) -> str:
    if n == 0:
        return "0 Б"
    for unit in ["Б", "КБ", "МБ", "ГБ", "ТБ"]:
        if abs(n) < 1024:
            return f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} ПБ"

# ========================== БАННЕР ==========================

BANNER = f"""
{c('╭───────────────────╮', Colors.DARK_RED)}\t{c('╭───────────────────────────────────────────────────────────────────────────────╮', Colors.DARK_RED)} {c('╭───────────────────╮', Colors.DARK_RED)}
{c('│', Colors.DARK_RED)}\t\t      {c('│', Colors.DARK_RED)}\t{c('│', Colors.DARK_RED)}    \t{c('▄▄▄', Colors.RED)}      {c('▄▄▄', Colors.DARK_RED)}  {c('▄▄▄', Colors.RED)}  {c('▄▄▄▄▄▄▄', Colors.DARK_RED)} {c('▄▄▄▄▄▄▄▄▄', Colors.RED)} {c('▄▄▄▄▄', Colors.DARK_RED)} {c('▄▄▄▄▄▄▄▄▄', Colors.RED)} {c('▄▄▄▄▄', Colors.DARK_RED)}   {c('▄▄▄▄', Colors.RED)}  \t{c('│', Colors.DARK_RED)} {c('│  by @frameworkq', Colors.GRAY)}   {c('│', Colors.DARK_RED)}
{c('│', Colors.DARK_RED)}\t\t      {c('│', Colors.DARK_RED)}\t{c('│', Colors.DARK_RED)}        {c('███', Colors.RED)}      {c('███', Colors.DARK_RED)}  {c('███', Colors.RED)} {c('█████▀▀▀', Colors.DARK_RED)} {c('▀▀▀███▀▀▀', Colors.RED)}  {c('███', Colors.DARK_RED)}  {c('▀▀▀███▀▀▀', Colors.RED)}  {c('███', Colors.DARK_RED)}  {c('▄██▀▀██▄', Colors.RED)}\t{c('│', Colors.DARK_RED)} {c('│  and @t1mott', Colors.GRAY)}      {c('│', Colors.DARK_RED)}
{c('│', Colors.DARK_RED)}\t\t      {c('│', Colors.DARK_RED)}\t{c('│', Colors.DARK_RED)}    \t{c('███', Colors.RED)}      {c('███', Colors.DARK_RED)}  {c('███', Colors.RED)}  {c('▀████▄', Colors.DARK_RED)}     {c('███', Colors.RED)}     {c('███', Colors.DARK_RED)}     {c('███', Colors.RED)}     {c('███', Colors.DARK_RED)}  {c('███', Colors.RED)}  {c('███', Colors.DARK_RED)} \t{c('│', Colors.DARK_RED)} {c('│', Colors.DARK_RED)}                   {c('│', Colors.DARK_RED)}
{c('│', Colors.DARK_RED)}\t\t      {c('│', Colors.DARK_RED)}\t{c('│', Colors.DARK_RED)}    \t {c('███', Colors.RED)}      {c('███▄▄███', Colors.DARK_RED)}    {c('▀████', Colors.RED)}    {c('███', Colors.DARK_RED)}     {c('███', Colors.RED)}     {c('███', Colors.DARK_RED)}  {c('███▀▀███', Colors.RED)} \t{c('│', Colors.DARK_RED)} {c('│  DataBase - 41TB', Colors.GRAY)}  {c('│', Colors.DARK_RED)}
{c('│', Colors.DARK_RED)}\t\t      {c('│', Colors.DARK_RED)}\t{c('│', Colors.DARK_RED)}       {c('████████', Colors.RED)} {c('▀██████▀', Colors.DARK_RED)} {c('███████▀', Colors.RED)}    {c('███', Colors.DARK_RED)}    {c('▄███▄', Colors.RED)}    {c('███', Colors.DARK_RED)}    {c('▄███▄', Colors.RED)} {c('███', Colors.DARK_RED)}  {c('███', Colors.RED)}\t{c('│', Colors.DARK_RED)} {c('│', Colors.DARK_RED)}                   {c('│', Colors.DARK_RED)}
{c('│', Colors.DARK_RED)}\t\t      {c('│', Colors.DARK_RED)}\t{c('│', Colors.DARK_RED)}    \t{c('│', Colors.DARK_RED)} {c('│', Colors.DARK_RED)}\t{c('│', Colors.DARK_RED)}    \t{c('│', Colors.DARK_RED)} {c('│  version - 1.0', Colors.GRAY)}   {c('│', Colors.DARK_RED)}
{c('╰───────────────────╯', Colors.DARK_RED)}\t{c('╰───────────────────────────────────────────────────────────────────────────────╯', Colors.DARK_RED)} {c('╰───────────────────╯', Colors.DARK_RED)}\t

{c('╭───────────────────────────────────────────────────────╮', Colors.DARK_RED)}
{c('│', Colors.DARK_RED)}  {c('[1]', Colors.RED)} - Поиск по ФИО      {c('│', Colors.DARK_RED)}  {c('[4]', Colors.RED)} - Поиск по ИНН        {c('│', Colors.DARK_RED)}
{c('│', Colors.DARK_RED)}  {c('[2]', Colors.RED)} - Поиск по номеру   {c('│', Colors.DARK_RED)}  {c('[5]', Colors.RED)} - Поиск по паспорту\t{c('│', Colors.DARK_RED)}        
{c('│', Colors.DARK_RED)}  {c('[3]', Colors.RED)} - Поиск по ном.авто {c('│', Colors.DARK_RED)}  {c('[6]', Colors.RED)} - Поиск по адресу\t{c('│', Colors.DARK_RED)}
{c('╰───────────────────────────────────────────────────────╯', Colors.DARK_RED)}
{c('╭───────────────────────────────────────────────────────╮', Colors.DARK_RED)}
{c('│', Colors.DARK_RED)}  {c('[7]', Colors.RED)} - Поиск по почте    {c('│', Colors.DARK_RED)}  {c('[10]', Colors.RED)} - Скоро...           {c('│', Colors.DARK_RED)}
{c('│', Colors.DARK_RED)}  {c('[8]', Colors.RED)} - Поиск по нику     {c('│', Colors.DARK_RED)}  {c('[11]', Colors.RED)} - Скоро...           {c('│', Colors.DARK_RED)}        
{c('│', Colors.DARK_RED)}  {c('[9]', Colors.RED)} - Multisearch       {c('│', Colors.DARK_RED)}  {c('[12]', Colors.RED)} - Я боюсь нато\t{c('│', Colors.DARK_RED)}
{c('│', Colors.DARK_RED)}  {c('[13]', Colors.RED)} - DB Uploader      {c('│', Colors.DARK_RED)}  {c('[14]', Colors.RED)} - Список баз         {c('│', Colors.DARK_RED)}
{c('╰───────────────────────────────────────────────────────╯', Colors.DARK_RED)}
"""

# ========================== DB UPLOADER API КЛИЕНТ ==========================

def load_config() -> dict:
    cfg = configparser.ConfigParser()
    if CONFIG_FILE.exists():
        cfg.read(CONFIG_FILE)
    return {
        "base_url": cfg.get("api", "base_url", fallback=BASE_URL),
        "api_key": cfg.get("api", "api_key", fallback=API_KEY),
    }

def save_config(base_url: str, api_key: str):
    cfg = configparser.ConfigParser()
    cfg["api"] = {"base_url": base_url.rstrip("/"), "api_key": api_key}
    with open(CONFIG_FILE, "w") as f:
        cfg.write(f)
    ok(f"Конфиг сохранён в {CONFIG_FILE}")

class DBClient:
    def __init__(self, base_url: str, api_key: str):
        self.base_url = base_url.rstrip("/")
        self.headers = {"x-api-key": api_key} if api_key else {}
        self.session = requests.Session()
        if self.headers:
            self.session.headers.update(self.headers)

    def _get(self, path: str, params=None, timeout=30) -> dict:
        url = f"{self.base_url}{path}"
        r = self.session.get(url, params=params, timeout=timeout)
        r.raise_for_status()
        return r.json()

    def _delete(self, path: str, timeout=15) -> dict:
        url = f"{self.base_url}{path}"
        r = self.session.delete(url, timeout=timeout)
        r.raise_for_status()
        return r.json()

    def _post(self, path: str, data=None, files=None, timeout=120) -> dict:
        url = f"{self.base_url}{path}"
        r = self.session.post(url, data=data, files=files, timeout=timeout)
        r.raise_for_status()
        return r.json()

    def list_databases(self) -> list:
        data = self._get("/databases")
        return data.get("databases", [])

    def get_database_info(self, name: str) -> dict:
        try:
            data = self._get(f"/databases/{requests.utils.quote(name, safe='')}")
            return data
        except requests.HTTPError as e:
            if e.response.status_code == 404:
                return {}
            raise

    def search_db(self, query: str, db: str = "", limit: int = 100) -> dict:
        params = {"q": query, "limit": limit}
        if db:
            params["db"] = db
        return self._get("/search", params=params, timeout=60)

    def delete_database(self, name: str) -> bool:
        try:
            self._delete(f"/databases/{requests.utils.quote(name, safe='')}")
            return True
        except requests.HTTPError as e:
            if e.response.status_code == 404:
                return False
            raise

    def upload(self, filepath: Path, name: str = "") -> dict:
        file_size = filepath.stat().st_size
        upload_name = name or filepath.stem[:100]

        print()
        info(f"Файл:  {filepath.name}")
        info(f"Размер: {fmt_size(file_size)}")
        info(f"База:  {upload_name}")
        hr()

        for attempt in range(1, MAX_RETRIES + 1):
            try:
                with open(filepath, "rb") as f:
                    with tqdm(
                        total=file_size,
                        unit="B",
                        unit_scale=True,
                        unit_divisor=1024,
                        desc=f"  Загрузка (попытка {attempt})",
                        colour="red",
                        bar_format="{desc}: {percentage:3.0f}% |{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]",
                    ) as bar:
                        class ProgressFile:
                            def __init__(self, fobj, pbar):
                                self._f = fobj
                                self._pbar = pbar
                            def read(self, size=-1):
                                chunk = self._f.read(size if size > 0 else CHUNK_SIZE)
                                self._pbar.update(len(chunk))
                                return chunk
                            def __len__(self):
                                return file_size

                        pf = ProgressFile(f, bar)
                        resp = self._post(
                            "/upload",
                            data={"name": upload_name},
                            files={"file": (filepath.name, pf)}
                        )

                return resp

            except requests.exceptions.ConnectionError:
                err(f"Нет соединения с сервером (попытка {attempt}/{MAX_RETRIES})")
            except requests.exceptions.Timeout:
                err(f"Таймаут (попытка {attempt}/{MAX_RETRIES})")
            except requests.exceptions.HTTPError as e:
                try:
                    detail = e.response.json().get("error", e.response.text)
                except Exception:
                    detail = str(e)
                err(f"HTTP {e.response.status_code}: {detail}")
                break
            except (KeyboardInterrupt, SystemExit):
                print()
                warn("Загрузка прервана")
                sys.exit(0)
            except Exception as e:
                err(f"Ошибка: {e} (попытка {attempt}/{MAX_RETRIES})")

            if attempt < MAX_RETRIES:
                print(Fore.YELLOW + f"  Повтор через {RETRY_DELAY} сек...")
                time.sleep(RETRY_DELAY)

        return {}

# ========================== УЛЬТРАБЫСТРЫЙ ПОИСК ==========================

def search(query: str, limit: int = 1000) -> dict:
    if not query.strip():
        return {"error": "Пустой запрос"}

    cache_key = f"{query.lower().strip()}_{limit}"
    if cache_key in cache:
        cache_time, cache_data = cache[cache_key]
        if time.time() - cache_time < CACHE_TTL:
            return cache_data

    params = {"q": query, "limit": limit}

    start_time = time.time()

    try:
        resp = requests.get(
            f"{BASE_URL}/search",
            params=params,
            timeout=30
        )

        elapsed_ms = round((time.time() - start_time) * 1000, 2)

        if resp.status_code == 200:
            data = resp.json()
            data["tookMs"] = elapsed_ms
            cache[cache_key] = (time.time(), data)
            return data
        elif resp.status_code == 401:
            return {"error": "❌ Неверный ключ API", "tookMs": elapsed_ms}
        elif resp.status_code == 404:
            return {"error": "❌ Эндпоинт не найден", "tookMs": elapsed_ms}
        else:
            return {"error": f"HTTP {resp.status_code}", "tookMs": elapsed_ms}

    except requests.exceptions.ConnectionError:
        return {"error": "❌ Сервер недоступен", "tookMs": 0}
    except requests.exceptions.Timeout:
        return {"error": "⏱️ Таймаут", "tookMs": 0}
    except Exception as e:
        return {"error": str(e), "tookMs": 0}

# ========================== ПОИСК ВСЕХ ЗАПИСЕЙ ==========================

def search_all(query: str, limit: int = 1000) -> dict:
    if not query.strip():
        return {"error": "Пустой запрос"}
    
    all_results = []
    page = 0
    page_size = 1000
    total_count = None
    
    start_time = time.time()
    
    print(c("\n⏳ Начинаю поиск всех записей...", Colors.YELLOW))
    
    try:
        while True:
            params = {
                "q": query, 
                "limit": page_size,
                "offset": page * page_size
            }
            
            resp = requests.get(
                f"{BASE_URL}/search",
                params=params,
                timeout=60
            )
            
            if resp.status_code != 200:
                if resp.status_code == 401:
                    return {"error": "❌ Неверный ключ API"}
                elif resp.status_code == 404:
                    return {"error": "❌ Эндпоинт не найден"}
                else:
                    return {"error": f"HTTP {resp.status_code}"}
            
            data = resp.json()
            results = data.get('results', [])
            
            if not results:
                break
            
            all_results.extend(results)
            
            if total_count is None:
                total_count = data.get('count', 0)
            
            print(c(f"  Загружено {len(all_results)} записей...", Colors.GRAY))
            
            if total_count and len(all_results) >= total_count:
                break
            
            if len(results) < page_size:
                break
            
            page += 1
            
            if len(all_results) > 1000000:
                break
        
        elapsed_ms = round((time.time() - start_time) * 1000, 2)
        
        print(c(f"✅ Найдено {len(all_results)} записей", Colors.GREEN))
        
        return {
            "results": all_results,
            "count": len(all_results),
            "totalCount": total_count,
            "tookMs": elapsed_ms
        }
        
    except requests.exceptions.ConnectionError:
        return {"error": "❌ Сервер недоступен", "tookMs": 0}
    except requests.exceptions.Timeout:
        return {"error": "⏱️ Таймаут", "tookMs": 0}
    except Exception as e:
        return {"error": str(e), "tookMs": 0}

def search_with_limit(query: str) -> dict:
    if not query.strip():
        return {"error": "Пустой запрос"}
    
    print(c("\n[?] Сколько записей загрузить?", Colors.CYAN))
    print(c("    [1] - 100 записей (быстро)", Colors.GRAY))
    print(c("    [2] - 1000 записей (стандарт)", Colors.GRAY))
    print(c("    [3] - 10000 записей (медленно)", Colors.GRAY))
    print(c("    [4] - ВСЕ записи (может быть очень медленно)", Colors.RED))
    
    choice = input(c("\n[+] Выберите опцию (1-4): ", Colors.RED)).strip()
    
    limit_map = {
        '1': 100,
        '2': 1000,
        '3': 10000,
        '4': None
    }
    
    limit = limit_map.get(choice, 1000)
    
    if limit is None:
        return search_all(query)
    else:
        return search(query, limit)

# ========================== JSON ВЫВОД ==========================

def print_json(data, query):
    clear()
    
    print("\n" + c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    print(c("  🔴 РЕЗУЛЬТАТ ПОИСКА", Colors.RED + Colors.BOLD))
    print(c("  Запрос: " + query, Colors.GRAY))
    
    if "tookMs" in data:
        speed_color = Colors.GREEN if data["tookMs"] < 10 else Colors.YELLOW if data["tookMs"] < 50 else Colors.RED
        print(c(f"  Время: {data['tookMs']} мс", speed_color))
    
    print(c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    
    if "error" in data:
        print(c(f"\n{data['error']}", Colors.RED))
        print(c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
        return
    
    results = data.get('results', [])
    total_count = data.get('totalCount', len(results))
    
    print(c(f"\n📊 Найдено: {total_count} записей", Colors.GREEN))
    print(c(f"📋 Загружено: {len(results)} записей", Colors.CYAN))
    
    if not results:
        print(c("\n❌ Ничего не найдено", Colors.RED))
        print(c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
        return
    
    show_all_results = results
    if len(results) > 100:
        show_all = input(c(f"\n[?] Показать все {len(results)} записей? (да/нет): ", Colors.YELLOW)).strip().lower()
        if show_all not in ['да', 'yes', 'y', 'д']:
            print(c("\n📋 Показаны первые 100 записей", Colors.CYAN))
            show_all_results = results[:100]
    
    print(c("\n📋 JSON:", Colors.CYAN))
    
    if len(results) > 500:
        print(c("[!] Слишком много записей для отображения. Сохраните результат в файл.", Colors.YELLOW))
    else:
        display_data = data.copy()
        display_data["results"] = show_all_results
        print(json.dumps(display_data, indent=2, ensure_ascii=False))
    
    print(c("\n" + "═" * 70, Colors.GRAY))
    print(c(f"  📊 ЗАПИСИ", Colors.GREEN))
    print(c("═" * 70, Colors.GRAY))
    
    for idx, item in enumerate(show_all_results, 1):
        print(c(f"\n[{idx}] ", Colors.DARK_RED + Colors.BOLD))
        if isinstance(item, dict):
            for key, val in item.items():
                if key.lower() in ['file', 'filename', 'database', 'db', 'databaseid', 'rank', 'id', 'datasetname', 'datasetid']:
                    continue
                if isinstance(val, dict):
                    print(f"  {c(key + ':', Colors.RED)}")
                    for k, v in val.items():
                        print(f"      {k}: {v}")
                else:
                    if val:
                        print(f"  {c(key + ':', Colors.RED)} {val}")
        else:
            print(f"  {item}")
        print(c("  " + "─" * 50, Colors.DARK_RED))
    
    print(c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    
    save = input(c("\n[+] Сохранить результат? (да/нет): ", Colors.GREEN)).strip().lower()
    if save in ['да', 'yes', 'y']:
        filename = f"search_{query}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(c(f"\n✅ Сохранено в: {filename}", Colors.GREEN))

# ========================== ФУНКЦИИ МЕНЮ ==========================

def show_menu():
    clear()
    print(BANNER)

def func_10():
    clear()
    print("\n" + c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    print(c("  🔴 ФУНКЦИЯ 10 — Скоро...", Colors.RED + Colors.BOLD))
    print(c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    print(c("\n[!] Эта функция будет доступна в следующей версии", Colors.YELLOW))
    input(c("\n[+] Нажмите Enter...", Colors.GRAY))

def func_11():
    clear()
    print("\n" + c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    print(c("  🔴 ФУНКЦИЯ 11 — Скоро...", Colors.RED + Colors.BOLD))
    print(c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    print(c("\n[!] Эта функция будет доступна в следующей версии", Colors.YELLOW))
    input(c("\n[+] Нажмите Enter...", Colors.GRAY))

def func_12():
    clear()
    print("\n" + c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    print(c("  🔴 Я БОЮСЬ НАТО", Colors.RED + Colors.BOLD))
    print(c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    print(c("\n[!] Нато — это Организация Североатлантического договора", Colors.YELLOW))
    print(c("[!] Шутка! Просто пасхалка", Colors.GRAY))
    print(c("[!] Автор: @frameworkq и @t1mott", Colors.GRAY))
    input(c("\n[+] Нажмите Enter...", Colors.GRAY))

def multisearch():
    clear()
    print("\n" + c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    print(c("  🔴 MULTISEARCH", Colors.RED + Colors.BOLD))
    print(c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    
    queries_input = input(c("\n[+] Введите запросы через запятую: ", Colors.RED)).strip()
    if not queries_input:
        return
    
    queries = [q.strip() for q in queries_input.split(',') if q.strip()]
    all_results = []
    
    for q in queries:
        print(c(f"\n[*] Поиск: {q}", Colors.CYAN))
        result = search(q)
        if result.get('results'):
            all_results.extend(result['results'])
    
    if all_results:
        print_json({"ok": True, "results": all_results, "count": len(all_results), "tookMs": 1}, "multisearch")
    else:
        print(c("\n❌ Ничего не найдено", Colors.RED))
    
    input(c("\n[+] Нажмите Enter...", Colors.GRAY))

# ========================== DB UPLOADER ФУНКЦИИ ==========================

def get_files_bulk(path_input: str) -> list:
    files = []
    
    if os.path.isdir(path_input):
        dir_path = Path(path_input)
        for f in dir_path.iterdir():
            if f.is_file() and f.suffix.lower() in SUPPORTED:
                files.append(f)
        return files
    
    if '*' in path_input or '?' in path_input:
        for pattern_path in glob.glob(path_input):
            p = Path(pattern_path)
            if p.is_file() and p.suffix.lower() in SUPPORTED:
                files.append(p)
        return files
    
    p = Path(path_input)
    if p.exists() and p.is_file() and p.suffix.lower() in SUPPORTED:
        return [p]
    
    return files

def db_uploader():
    cfg = load_config()
    
    if not cfg["base_url"]:
        cfg["base_url"] = BASE_URL
    
    client = DBClient(cfg["base_url"], cfg.get("api_key", ""))
    
    clear()
    print("\n" + c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    print(c("  🔴 DB UPLOADER — Пакетная загрузка баз", Colors.RED + Colors.BOLD))
    print(c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    
    print(c("\n📌 Введите путь:", Colors.CYAN))
    print(c("  • Папка: /home/user/data/          - загрузит ВСЕ файлы", Colors.GRAY))
    print(c("  • Шаблон: *.csv                    - все CSV файлы", Colors.GRAY))
    print(c("  • Файл: data.csv                   - конкретный файл", Colors.GRAY))
    print(c(f"  • Поддерживаемые форматы: {', '.join(sorted(SUPPORTED))}", Colors.GRAY))
    
    path_input = input(c("\n[+] Путь: ", Colors.RED)).strip()
    
    if not path_input:
        print(c("\n❌ Путь не указан", Colors.YELLOW))
        input(c("\n[+] Нажмите Enter...", Colors.GRAY))
        return
    
    files = get_files_bulk(path_input)
    
    if not files:
        print(c(f"\n❌ Файлы не найдены: {path_input}", Colors.YELLOW))
        input(c("\n[+] Нажмите Enter...", Colors.GRAY))
        return
    
    print(c(f"\n📁 Найдено файлов: {len(files)}", Colors.GREEN))
    for f in files[:10]:
        size = fmt_size(f.stat().st_size)
        print(c(f"  • {f.name} ({size})", Colors.WHITE))
    if len(files) > 10:
        print(c(f"  ... и ещё {len(files) - 10} файлов", Colors.GRAY))
    
    total_size = sum(f.stat().st_size for f in files)
    print(c(f"\n💾 Общий размер: {fmt_size(total_size)}", Colors.CYAN))
    
    confirm = input(c("\n[+] Начать ПАКЕТНУЮ загрузку ВСЕХ файлов? (да/нет): ", Colors.RED)).strip().lower()
    if confirm not in ['да', 'yes', 'y', 'д']:
        print(c("\n❌ Загрузка отменена", Colors.YELLOW))
        input(c("\n[+] Нажмите Enter...", Colors.GRAY))
        return
    
    print()
    hr()
    print(c("  ⏳ ЗАГРУЗКА {} ФАЙЛОВ...".format(len(files)), Colors.RED + Colors.BOLD))
    hr()
    
    success_count = 0
    failed = []
    total_rows = 0
    total_uploaded_size = 0
    
    for i, filepath in enumerate(files, 1):
        print()
        print(c(f"[{i}/{len(files)}] {filepath.name}", Colors.CYAN + Colors.BOLD))
        
        db_name = filepath.stem[:100]
        
        result = client.upload(filepath, db_name)
        
        if result.get("ok"):
            db = result.get("database", {})
            rows = db.get('rowCount', 0)
            total_rows += rows
            total_uploaded_size += filepath.stat().st_size
            print(c(f"  ✅ {db_name} ({rows:,} строк)", Colors.GREEN))
            success_count += 1
        else:
            error = result.get('error', 'Ошибка')
            print(c(f"  ❌ {error}", Colors.RED))
            failed.append(filepath.name)
    
    print()
    hr()
    print(c("  📊 ИТОГ ПАКЕТНОЙ ЗАГРУЗКИ", Colors.RED + Colors.BOLD))
    hr()
    print(c(f"  ✅ Успешно загружено: {success_count}/{len(files)}", Colors.GREEN))
    print(c(f"  📊 Всего строк: {total_rows:,}", Colors.CYAN))
    print(c(f"  💾 Загружено данных: {fmt_size(total_uploaded_size)}", Colors.CYAN))
    
    if failed:
        print(c(f"\n  ❌ Не удалось загрузить:", Colors.RED))
        for f in failed:
            print(c(f"    • {f}", Colors.RED))
    
    input(c("\n[+] Нажмите Enter...", Colors.GRAY))

def db_list():
    cfg = load_config()
    if not cfg["base_url"]:
        cfg["base_url"] = BASE_URL
    
    client = DBClient(cfg["base_url"], cfg.get("api_key", ""))
    
    clear()
    print("\n" + c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    print(c("  📋 СПИСОК БАЗ ДАННЫХ", Colors.RED + Colors.BOLD))
    print(c("═" * 70, Colors.BLOOD_RED + Colors.BOLD))
    
    try:
        dbs = client.list_databases()
    except Exception as e:
        print(c(f"\n❌ Ошибка: {e}", Colors.RED))
        input(c("\n[+] Нажмите Enter...", Colors.GRAY))
        return
    
    if not dbs:
        print(c("\n❌ Нет загруженных баз", Colors.YELLOW))
    else:
        total_size = 0
        
        for i, db in enumerate(dbs, 1):
            db_name = db.get("name", "—")
            created = db.get("createdAt", "")[:10]
            cols = db.get("columns", [])
            row_count = db.get("rowCount", 0)
            
            db_size = 0
            try:
                db_info = client.get_database_info(db_name)
                if db_info:
                    db_size = db_info.get("size", 0)
                    if not db_size:
                        db_size = db_info.get("sizeBytes", 0)
                    if not db_size:
                        db_size = db_info.get("fileSize", 0)
            except Exception:
                db_size = row_count * 500 + len(cols) * 100
            
            total_size += db_size
            
            col_preview = ", ".join(cols[:5]) + (f" +{len(cols)-5}" if len(cols) > 5 else "")
            
            print(c(f"\n[{i}] ", Colors.DARK_RED + Colors.BOLD) + c(db_name, Colors.WHITE + Colors.BOLD))
            print(c(f"    📊 {row_count:,} строк  |  📅 {created}", Colors.GRAY))
            print(c(f"    🗂  {col_preview}", Colors.GRAY))
            
            size_str = fmt_size(db_size)
            if db_size > 100 * 1024 * 1024:
                size_color = Colors.RED
            elif db_size > 10 * 1024 * 1024:
                size_color = Colors.YELLOW
            else:
                size_color = Colors.GREEN
            
            print(c(f"    💾 {size_str}", size_color))
        
        print(c(f"\n📊 Всего: {len(dbs)} баз", Colors.GREEN))
        print(c(f"💾 Общий размер: {fmt_size(total_size)}", Colors.CYAN))
    
    input(c("\n[+] Нажмите Enter...", Colors.GRAY))

# ========================== ОСНОВНАЯ ФУНКЦИЯ ==========================

def main():
    # Автозапуск сервера в фоне
    print(c("\n⏳ Запуск API сервера...", Colors.YELLOW))
    if ensure_server():
        print(c("✅ Сервер запущен на " + BASE_URL, Colors.GREEN))
    else:
        print(c("❌ Не удалось запустить сервер!", Colors.RED))
        print(c("   Установи зависимости: pip install fastapi uvicorn python-multipart openpyxl", Colors.GRAY))
        sys.exit(1)
    time.sleep(0.5)

    labels = {
        '1': 'ФИО', '2': 'номер телефона', '3': 'номер автомобиля',
        '4': 'ИНН', '5': 'паспорт', '6': 'адрес', '7': 'почту', '8': 'никнейм'
    }
    
    while True:
        show_menu()
        choice = input(c("\n[+] Выберите действие: ", Colors.RED)).strip()
        
        if choice == '0':
            print(c("\n[!] Выход", Colors.RED))
            sys.exit(0)
        
        if choice in labels:
            query = input(c(f"\n[+] Введите {labels[choice]}: ", Colors.RED)).strip()
            if query:
                result = search_with_limit(query)
                print_json(result, query)
            else:
                print(c("\n[!] Пустой запрос", Colors.YELLOW))
            input(c("\n[+] Нажмите Enter...", Colors.GRAY))
        
        elif choice == '9':
            multisearch()
        
        elif choice == '10':
            func_10()
        elif choice == '11':
            func_11()
        elif choice == '12':
            func_12()
        elif choice == '13':
            db_uploader()
        elif choice == '14':
            db_list()
        
        else:
            print(c("[!] Неверный выбор", Colors.RED))
            input(c("\n[+] Нажмите Enter...", Colors.GRAY))

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print(c("\n[!] Выход", Colors.RED))
        sys.exit(0)
