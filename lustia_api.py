"""Lustia API Search — сервер для загрузки баз данных любого формата и полнотекстового поиска.

Поддерживаемые форматы: CSV, TSV, JSON, XML, Excel (.xlsx/.xls), TXT

Запуск:
    pip install fastapi uvicorn python-multipart openpyxl
    python lustia_api.py

Документация: http://localhost:8000/docs
"""

import csv
import io
import json
import sqlite3
import threading
import time
import xml.etree.ElementTree as ET
from contextlib import asynccontextmanager
from pathlib import Path

import uvicorn
from fastapi import FastAPI, APIRouter, Form, HTTPException, Query, UploadFile
from openpyxl import load_workbook


# ──────────────────────────────────────────────
# DATABASE (SQLite + FTS5)
# ──────────────────────────────────────────────

_DB_DIR = Path(__file__).resolve().parent / "data"
_DB_PATH = _DB_DIR / "store.db"
_local = threading.local()


def _get_conn() -> sqlite3.Connection:
    conn = getattr(_local, "conn", None)
    if conn is None:
        _DB_DIR.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(str(_DB_PATH), check_same_thread=False)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        _local.conn = conn
    return conn


def init_db() -> None:
    conn = _get_conn()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS databases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            db_id INTEGER NOT NULL REFERENCES databases(id) ON DELETE CASCADE,
            row_data TEXT NOT NULL
        );
        CREATE VIRTUAL TABLE IF NOT EXISTS records_fts USING fts5(
            row_data, content='records', content_rowid='id'
        );
        CREATE TRIGGER IF NOT EXISTS records_ai AFTER INSERT ON records BEGIN
            INSERT INTO records_fts(rowid, row_data) VALUES (new.id, new.row_data);
        END;
        CREATE TRIGGER IF NOT EXISTS records_ad AFTER DELETE ON records BEGIN
            INSERT INTO records_fts(records_fts, rowid, row_data)
                VALUES ('delete', old.id, old.row_data);
        END;
    """)
    conn.commit()


def create_database_entry(name: str) -> int:
    conn = _get_conn()
    cur = conn.execute("INSERT INTO databases (name) VALUES (?)", (name,))
    conn.commit()
    return cur.lastrowid


def get_database_entry(name: str) -> dict | None:
    conn = _get_conn()
    row = conn.execute("SELECT * FROM databases WHERE name = ?", (name,)).fetchone()
    return dict(row) if row else None


def list_databases_db() -> list[dict]:
    conn = _get_conn()
    rows = conn.execute("SELECT * FROM databases ORDER BY id").fetchall()
    return [dict(r) for r in rows]


def delete_database_entry(db_id: int) -> None:
    conn = _get_conn()
    conn.execute("DELETE FROM records WHERE db_id = ?", (db_id,))
    conn.execute("DELETE FROM databases WHERE id = ?", (db_id,))
    conn.commit()


def insert_records(db_id: int, rows: list[str]) -> int:
    conn = _get_conn()
    conn.executemany(
        "INSERT INTO records (db_id, row_data) VALUES (?, ?)",
        [(db_id, r) for r in rows],
    )
    conn.commit()
    return len(rows)


def get_columns(db_id: int) -> list[str]:
    """Extract column names from the first record of a database."""
    conn = _get_conn()
    row = conn.execute(
        "SELECT row_data FROM records WHERE db_id = ? LIMIT 1", (db_id,)
    ).fetchone()
    if not row:
        return []
    parts = row["row_data"].split(" | ")
    cols = []
    for p in parts:
        if ": " in p:
            cols.append(p.split(": ", 1)[0])
    return cols


def search_records(
    query: str, limit: int = 50, offset: int = 0, db_name: str | None = None
) -> tuple[list[dict], int]:
    conn = _get_conn()

    # Count total matches
    if db_name:
        count_sql = """
            SELECT COUNT(*) AS cnt
            FROM records_fts fts
            JOIN records r ON r.id = fts.rowid
            JOIN databases d ON d.id = r.db_id
            WHERE fts.row_data MATCH ? AND d.name = ?
        """
        total = conn.execute(count_sql, (query, db_name)).fetchone()["cnt"]
        sql = """
            SELECT r.id, d.name AS db_name, r.row_data, rank
            FROM records_fts fts
            JOIN records r ON r.id = fts.rowid
            JOIN databases d ON d.id = r.db_id
            WHERE fts.row_data MATCH ? AND d.name = ?
            ORDER BY rank LIMIT ? OFFSET ?
        """
        rows = conn.execute(sql, (query, db_name, limit, offset)).fetchall()
    else:
        count_sql = """
            SELECT COUNT(*) AS cnt
            FROM records_fts fts
            JOIN records r ON r.id = fts.rowid
            WHERE fts.row_data MATCH ?
        """
        total = conn.execute(count_sql, (query,)).fetchone()["cnt"]
        sql = """
            SELECT r.id, d.name AS db_name, r.row_data, rank
            FROM records_fts fts
            JOIN records r ON r.id = fts.rowid
            JOIN databases d ON d.id = r.db_id
            WHERE fts.row_data MATCH ?
            ORDER BY rank LIMIT ? OFFSET ?
        """
        rows = conn.execute(sql, (query, limit, offset)).fetchall()

    results = []
    for r in rows:
        item: dict = {"database": r["db_name"]}
        for part in r["row_data"].split(" | "):
            if ": " in part:
                k, v = part.split(": ", 1)
                item[k] = v
            else:
                item["data"] = part
        results.append(item)

    return results, total


def count_records(db_id: int) -> int:
    conn = _get_conn()
    row = conn.execute(
        "SELECT COUNT(*) AS cnt FROM records WHERE db_id = ?", (db_id,)
    ).fetchone()
    return row["cnt"]


# ──────────────────────────────────────────────
# PARSER (все форматы)
# ──────────────────────────────────────────────

_ENCODINGS = ["utf-8", "utf-8-sig", "cp1251", "latin-1"]


def _decode(content: bytes) -> str:
    for enc in _ENCODINGS:
        try:
            return content.decode(enc)
        except (UnicodeDecodeError, UnicodeError):
            continue
    return content.decode("latin-1")


def parse_csv(content: bytes) -> list[str]:
    text = _decode(content)
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for record in reader:
        row_text = " | ".join(f"{k}: {v}" for k, v in record.items() if v)
        if row_text:
            rows.append(row_text)
    return rows


def parse_tsv(content: bytes) -> list[str]:
    text = _decode(content)
    reader = csv.DictReader(io.StringIO(text), delimiter="\t")
    rows = []
    for record in reader:
        row_text = " | ".join(f"{k}: {v}" for k, v in record.items() if v)
        if row_text:
            rows.append(row_text)
    return rows


def parse_json(content: bytes) -> list[str]:
    text = _decode(content)
    data = json.loads(text)
    if isinstance(data, list):
        items = data
    elif isinstance(data, dict):
        items = [data]
    else:
        raise ValueError(f"Unsupported JSON root type: {type(data).__name__}")
    rows = []
    for item in items:
        if isinstance(item, dict):
            row_text = " | ".join(
                f"{k}: {v}" for k, v in item.items() if v is not None
            )
        else:
            row_text = str(item)
        if row_text:
            rows.append(row_text)
    return rows


def parse_xml(content: bytes) -> list[str]:
    root = ET.fromstring(content)
    rows = []
    for element in root:
        parts = []
        if element.text and element.text.strip():
            parts.append(f"{element.tag}: {element.text.strip()}")
        for child in element:
            text = child.text.strip() if child.text else ""
            if text:
                parts.append(f"{child.tag}: {text}")
            for attr_key, attr_val in child.attrib.items():
                parts.append(f"{attr_key}: {attr_val}")
        for attr_key, attr_val in element.attrib.items():
            parts.append(f"{attr_key}: {attr_val}")
        if parts:
            rows.append(" | ".join(parts))
    return rows


def parse_excel(content: bytes) -> list[str]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_rows = list(ws.iter_rows(values_only=True))
        if not sheet_rows:
            continue
        headers = [
            str(h) if h is not None else f"col_{i}"
            for i, h in enumerate(sheet_rows[0])
        ]
        for data_row in sheet_rows[1:]:
            parts = []
            for header, val in zip(headers, data_row):
                if val is not None and str(val).strip():
                    parts.append(f"{header}: {val}")
            if parts:
                rows.append(" | ".join(parts))
    wb.close()
    return rows


def parse_txt(content: bytes) -> list[str]:
    text = _decode(content)
    rows = []
    for line in text.splitlines():
        stripped = line.strip()
        if stripped:
            rows.append(stripped)
    return rows


SUPPORTED_EXTENSIONS = {
    ".csv": "CSV",
    ".tsv": "TSV",
    ".json": "JSON",
    ".xml": "XML",
    ".xlsx": "Excel",
    ".xls": "Excel",
    ".txt": "Text",
}


def parse_file(filename: str, content: bytes) -> list[str]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return parse_csv(content)
    if lower.endswith(".tsv"):
        return parse_tsv(content)
    if lower.endswith(".json"):
        return parse_json(content)
    if lower.endswith(".xml"):
        return parse_xml(content)
    if lower.endswith((".xlsx", ".xls")):
        return parse_excel(content)
    if lower.endswith(".txt"):
        return parse_txt(content)
    supported = ", ".join(SUPPORTED_EXTENSIONS.keys())
    raise ValueError(f"Unsupported file type: {filename}. Supported: {supported}")


# ──────────────────────────────────────────────
# API ROUTES
# ──────────────────────────────────────────────

router = APIRouter()


@router.post("/upload")
async def upload_database(file: UploadFile, name: str = Form(...)):
    """Загрузить файл как новую базу данных."""
    start = time.perf_counter()

    if not file.filename:
        raise HTTPException(status_code=400, detail="File has no name")

    db_name = name.strip()[:100]
    if not db_name:
        raise HTTPException(status_code=400, detail="Empty database name")

    existing = get_database_entry(db_name)
    if existing:
        raise HTTPException(
            status_code=409, detail=f"Database '{db_name}' already exists"
        )

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        rows = parse_file(file.filename, content)
    except (ValueError, UnicodeDecodeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    if not rows:
        raise HTTPException(status_code=400, detail="File contains no records")

    db_id = create_database_entry(db_name)
    inserted = insert_records(db_id, rows)
    elapsed = time.perf_counter() - start

    columns = get_columns(db_id)

    return {
        "ok": True,
        "database": {
            "name": db_name,
            "rowCount": inserted,
            "columns": columns,
            "createdAt": get_database_entry(db_name)["created_at"],
        },
        "elapsed_seconds": round(elapsed, 4),
    }


@router.get("/databases")
async def list_all_databases():
    """Список всех загруженных баз данных."""
    dbs = list_databases_db()
    result = []
    for db in dbs:
        cnt = count_records(db["id"])
        columns = get_columns(db["id"])
        result.append(
            {
                "name": db["name"],
                "rowCount": cnt,
                "columns": columns,
                "createdAt": db["created_at"],
            }
        )
    return {"databases": result}


@router.get("/databases/{db_name}")
async def get_database_info(db_name: str):
    """Информация о конкретной базе данных."""
    entry = get_database_entry(db_name)
    if not entry:
        raise HTTPException(
            status_code=404, detail=f"Database '{db_name}' not found"
        )
    cnt = count_records(entry["id"])
    columns = get_columns(entry["id"])
    return {
        "name": entry["name"],
        "rowCount": cnt,
        "columns": columns,
        "createdAt": entry["created_at"],
    }


@router.delete("/databases/{db_name}")
async def delete_database(db_name: str):
    """Удалить базу данных и все её записи."""
    entry = get_database_entry(db_name)
    if not entry:
        raise HTTPException(
            status_code=404, detail=f"Database '{db_name}' not found"
        )
    delete_database_entry(entry["id"])
    return {"ok": True, "deleted": db_name}


@router.get("/formats")
async def supported_formats():
    """Список поддерживаемых форматов файлов."""
    return {"formats": SUPPORTED_EXTENSIONS}


@router.get("/search")
async def search_endpoint(
    q: str = Query(..., min_length=1, description="Search query"),
    db: str | None = Query(None, description="Limit to specific database"),
    limit: int = Query(100, ge=1, le=10000),
    offset: int = Query(0, ge=0),
):
    """Полнотекстовый поиск по всем базам данных."""
    start = time.perf_counter()

    if db:
        entry = get_database_entry(db)
        if not entry:
            raise HTTPException(
                status_code=404, detail=f"Database '{db}' not found"
            )

    try:
        results, total_count = search_records(
            q, limit=limit, offset=offset, db_name=db
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Search error: {exc}")

    elapsed_ms = round((time.perf_counter() - start) * 1000, 2)

    return {
        "ok": True,
        "results": results,
        "count": total_count,
        "totalCount": total_count,
        "tookMs": elapsed_ms,
    }


# ──────────────────────────────────────────────
# APP
# ──────────────────────────────────────────────


@asynccontextmanager
async def lifespan(a: FastAPI):
    init_db()
    yield


app = FastAPI(
    title="Lustia API Search",
    description="Lustia API — загрузка и поиск по базам данных любого формата за <1 секунду",
    version="1.0.0",
    lifespan=lifespan,
)

app.include_router(router, prefix="/api")


@app.get("/health")
async def health():
    return {"status": "ok"}


if __name__ == "__main__":
    uvicorn.run(
        "lustia_api:app",
        host="0.0.0.0",
        port=8000,
        reload=True,
        reload_includes=["lustia_api.py"],
    )
